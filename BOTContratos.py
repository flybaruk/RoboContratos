import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
from datetime import datetime
import time
import pyautogui
import os
from pathlib import Path
import pandas as pd
import numpy as np
import locale
import glob
import shutil
from openpyxl import workbook
import xlrd
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Definir a localização como "pt_BR" (Português do Brasil)
locale.setlocale(locale.LC_ALL, 'pt_BR')

user_dir = Path.home()# Obtém o diretório base do usuário atual

downloads_dir = os.path.join(user_dir, "Downloads")# Concatena o diretório de downloads com o diretório base do usuário

# Configurar as opções do Chrome para abrir maximizado
chrome_options = Options()
chrome_options.add_argument("--start-maximized") #testar o argumento  '--headless=new' para a automação rodar em segundo plano
chrome_options.add_argument("--headless=new") #segundo plano

nome_arquivo = 'Planilha Contratos.xlsx'

#Definindo as configurações do chrome
driver = webdriver.Chrome(options=chrome_options)
acaodriver = ActionChains(driver)
acao = ActionChains(driver)

for arquivo in os.listdir(downloads_dir):
    if arquivo == nome_arquivo:

        caminho_arquivo = os.path.join(downloads_dir, arquivo)
    
        os.remove(caminho_arquivo)

        print('O arquivo foi apagado com sucesso')

        break
else:
    print('O arquivo não foi encontrado')


nome_arquivo2 = 'Planilha Contratos.xls'

for arquivo2 in os.listdir(downloads_dir):
    if arquivo2 == nome_arquivo2:

        caminho_arquivo2 = os.path.join(downloads_dir, arquivo2)
    
        os.remove(caminho_arquivo2)

        print('O arquivo foi apagado com sucesso')

        break
else:
    print('O arquivo não foi encontrado')

email_usuario = "famasilva@tjba.jus.br"
senha = "Fabio@tjba2024"

#entrando no site
driver.get("https://tjba.contratosgov.com.br/App#")

#digitando usuario e senha
driver.find_element(By.ID, "Usuarios_dsLogin").send_keys(email_usuario)
driver.find_element(By.ID, "Usuarios_dsSenha").send_keys(senha)
time.sleep(2)

#botao entrar
driver.find_element(By.XPATH, "/html/body/main/div/div/div/form[1]/div/input").click()

#acessando menu relatorios
btn_relatorios = WebDriverWait(driver, 35).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/section/div[1]/div[1]/div[7]")))
btn_relatorios.click()

#acessando submenu relatorios gerais
btn_gerais = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/section/div[1]/div[2]/div[1]/div[1]")))
btn_gerais.click()

time.sleep(2)

# Localizando o iframe
iframe = driver.find_element(By.ID, "frame_conteudo_1")
driver.switch_to.frame(iframe)

# Identificando o drilldown
tipo_drilldown = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "tp_Listagem")))

# Cria o elemento que vai interagir com o drilldown
select = Select(tipo_drilldown)

# Selecionando a opção pelo indice
select.select_by_index(2)

#segundo drilldown
nome_relatorio = WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.ID, "tp_Relatorio")))

#seleciona pelo indice
select2 = Select(nome_relatorio)
select2.select_by_index(1)

#aperta em gerar relatório
btn_gerar = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "/html/body/div[2]/div/div[3]/div/div/button[2]")))
btn_gerar.click()
time.sleep(5)

#saindo do iframe
driver.switch_to.default_content()

pyautogui.press('tab', presses=2)
pyautogui.hotkey('enter')
time.sleep(0.5)
pyautogui.press('tab', presses=1)      
pyautogui.hotkey('enter')
time.sleep(0.5)
pyautogui.press('tab', presses=3)
pyautogui.hotkey('enter')
time.sleep(0.5)
pyautogui.press('tab', presses=2)
pyautogui.hotkey('enter')

time.sleep(3)


#Renomeando o arquivo
caminho_download = r'c:\Users\lhgarces\Downloads'
arquivos = os.listdir(caminho_download)
ultimo_download = max(arquivos, key=lambda x: os.path.getctime(os.path.join(caminho_download, x)))
novo_nome = "Planilha Contratos.xls"
os.rename(os.path.join(caminho_download, ultimo_download), os.path.join(caminho_download, novo_nome))


input_path = r'C:\Users\lhgarces\Downloads\Planilha Contratos.xls'
output_path = r'C:\Users\lhgarces\Downloads\Planilha Contratos.xlsx'

# Carregue o arquivo Excel usando xlrd e pandas
workbook_xlrd = xlrd.open_workbook(input_path, ignore_workbook_corruption=True)
df = pd.read_excel(workbook_xlrd, header = 0)

df = df.iloc[1::2, :]
print(df)

# Crie um novo arquivo Excel (.xlsx) usando openpyxl
workbook_openpyxl = openpyxl.Workbook()
sheet = workbook_openpyxl.active

# Preencha o novo arquivo com os dados do DataFrame
for r_idx, row in enumerate(df.values, 1):
    for c_idx, value in enumerate(row, 1):
        sheet.cell(row=r_idx + 1, column=c_idx, value=value)

# Adicionar cabeçalho
for c_idx, col_name in enumerate(df.columns, 1):
    sheet.cell(row=1, column=c_idx, value=col_name)

    
    # Crie uma tabela
table = Table(displayName="SuaTabela", ref=sheet.dimensions)

#Renomeando a aba
sheet.title = "Worksheet"

# Defina o estilo da tabela (opcional)
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style

# Adicione a tabela à planilha
sheet.add_table(table)

# Salve as alterações no arquivo Excel
workbook_openpyxl.save(output_path)

driver.quit()


driver2 = webdriver.Chrome(options=chrome_options)
acaodriver2 = ActionChains(driver2)
acao2 = ActionChains(driver2)
wait2 = WebDriverWait(driver2, 10)

#entrando no sharepoint
usuario_outlook = "jvafreitas@tjba.jus.br"
senha_outlook = "joao@tjba2023"
driver2.get('https://tjbacotec.sharepoint.com/sites/CGTIC')

time.sleep(2.5)

#digitando o login
#usuario
driver2.find_element(By.ID, 'i0116').send_keys(usuario_outlook)

#avançar
btn_avancar = wait2.until(EC.presence_of_element_located((By.XPATH,'/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div[1]/div[3]/div/div/div/div[4]/div/div/div/div/input')))
btn_avancar.click()

time.sleep(1.5)

#senha
driver2.find_element(By.ID, 'i0118').send_keys(senha_outlook)

#clicando entrar
btn_entrar = wait2.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div/form[1]/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[5]/div/div/div/div/input')))
btn_entrar.click()

#permanecer conectado
btn_sim = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div/form/div/div/div[2]/div[1]/div/div/div/div/div/div[3]/div/div[2]/div/div[3]/div[2]/div/div/div[2]/input')))
btn_sim.click()

#Clica em configurações da pagina
settings = wait2.until(EC.element_to_be_clickable((By.XPATH,'/html/body/div/div[1]/div/div[1]/div[2]/div/div/div/div/div/div[3]/div[6]/div/button/div/span')))
settings.click()

#Clica em conteúdo do site
ctd_site = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[2]/div[2]/div[3]/div/div/div/div/div[1]/div/div[2]/ul/li[3]/a')))
ctd_site.click()

#Entra na pasta documentos
documentos = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[3]/div[2]/div[2]/div[2]/div[3]/div[2]/div[2]/div[1]/div/div/div/div/div[3]/div/div[2]/div/div/div/div/div/div/div/div[2]/div/div/div/div/div[1]/div[4]/div/div/div/div[2]/div/div/a')))
documentos.click()

#Clica nos 3 pontinhos da planilha chamados
tp = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[3]/div/div[2]/div[2]/div[2]/div[2]/div[1]/div/div/div/div/div/div[3]/div[2]/div/div/div/div[2]/div/div/div/div/div/div[6]/div/div/div[2]/div[2]/div/div[1]/div[1]/span/span/button')))
acao2.context_click(tp).perform()

time.sleep(2)
#Apaga a planilha atual
eliminar_planilha = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/div/div/div/ul/li[7]/button/div/span')))
eliminar_planilha.click()

time.sleep(2)

#confirmar exclusão
excluir = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/div[2]/div[2]/div/div[2]/div[2]/div/span[1]/button/span/span/span')))
excluir.click()

time.sleep(2)

#Carrega a nova planilha
carregar = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[3]/div/div[2]/div[2]/div[2]/div[1]/div/div/div/div/div/div/div[1]/div[2]/button/span')))
carregar.click()

time.sleep(2)

#Abre os ficheiros
ficheiros = wait2.until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[3]/div/div/div/div/div/div/ul/li[1]/button')))
ficheiros.click()

time.sleep(1.5)

#Seleciona o arquivo
pyautogui.write(r"C:\Users\lhgarces\Downloads\Planilha Contratos.xlsx")
pyautogui.hotkey("enter")

time.sleep(15)
driver2.quit()